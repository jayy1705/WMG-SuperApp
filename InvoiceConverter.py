"""
InvoiceConverter.py

Conversion logic for WMG's invoice listing .xlsx files, with no GUI.
Same pattern as cleaning.py: every function here takes a file path or
DataFrame in and returns data out — no Tkinter, no file dialogs, no
print statements.

The GUI screen that calls this lives in gui/step1_convert.py.
"""

import re
import zipfile
import tempfile
import datetime
import os
from collections import defaultdict

import openpyxl
import pandas as pd

FIELDS = [
    "DocNo", "DocDate", "Code", "Name", "InvoiceAmount",
    "Seq", "GLCode", "Description", "Project", "Quantity",
    "UOM", "UnitPrice", "LineAmount",
]


def patch_xlsx(input_path: str) -> str:
    """Fix known non-standard XML attributes so openpyxl can open the file."""
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)

    with zipfile.ZipFile(input_path, "r") as zin:
        names = zin.namelist()
        wb_xml = zin.read("xl/workbook.xml").decode("utf-8") if "xl/workbook.xml" in names else None
        sheet_xmls = {
            n: zin.read(n).decode("utf-8")
            for n in names
            if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")
        }

        if wb_xml is not None:
            wb_xml = wb_xml.replace("WindowWidth", "windowWidth").replace("WindowHeight", "windowHeight")

        for n in list(sheet_xmls.keys()):
            sheet_xmls[n] = re.sub(r'firstPageNo="[^"]*"\s*', "", sheet_xmls[n])

        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                if wb_xml is not None and item.filename == "xl/workbook.xml":
                    content = wb_xml.encode("utf-8")
                elif item.filename in sheet_xmls:
                    content = sheet_xmls[item.filename].encode("utf-8")
                zout.writestr(item, content)

    return tmp_path


def is_blank(row):
    return all(c is None for c in row)


def flatten(rows):
    """Walks the raw sheet rows and returns (records, unrecognized_rows)."""
    records = []
    cur_inv = None
    cur_line = None
    unrecognized = []

    for i, row in enumerate(rows):
        # Only columns up to c15 carry data this parser uses (c14 is
        # skipped but kept positionally so the names stay column-accurate).
        c = (list(row) + [None] * 16)[:16]
        (c0, c1, c2, c3, c4, c5, c6, c7, c8, c9,
         c10, c11, c12, c13, _, c15) = c

        if c8 == "Grand Total Amount (RM)":
            break

        if is_blank(row):
            continue

        if isinstance(c0, str) and "WEST MALAYAN GROUP" in c0:
            continue

        if c0 in ("Doc. No", "Seq"):
            continue

        if c1 == " : ":
            continue
        if isinstance(c0, str) and c0.strip() in ("Item", "Category"):
            continue

        if c7 == "Invoice Listing" or (isinstance(c7, str) and c7.startswith("As At")):
            continue

        if isinstance(c0, str) and isinstance(c2, datetime.datetime) and isinstance(c15, (int, float)):
            cur_inv = {
                "DocNo": c0,
                "DocDate": c2.strftime("%Y-%m-%d"),
                "Code": c4,
                "Name": c6,
                "InvoiceAmount": c15,
            }
            cur_line = None
            continue

        if (
            c0 is None
            and c6
            and all(v is None for v in [c1, c2, c3, c4, c5, c7, c8, c9, c10, c11, c12, c13])
            and cur_inv is not None
            and cur_line is None
        ):
            cur_inv["Name"] = ((cur_inv["Name"] or "") + " " + str(c6)).strip()
            continue

        if isinstance(c0, (int, float)) and cur_inv is not None:
            cur_line = {
                **cur_inv,
                "Seq": c0,
                "GLCode": c1,
                "Description": c3 or "",
                "Project": c8,
                "Quantity": c10,
                "UOM": c11,
                "UnitPrice": c12,
                "LineAmount": c13,
            }
            records.append(cur_line)
            continue

        if c0 is None and c3 and cur_line is not None:
            cur_line["Description"] = (cur_line["Description"] + " " + str(c3)).strip()
            continue

        unrecognized.append((i, row))

    return records, unrecognized


def convert_file(input_path: str):
    """
    Runs the full conversion for a given input path.
    Returns (records, unrecognized, mismatches, invoice_count).
    """
    patched_path = patch_xlsx(input_path)
    try:
        wb = openpyxl.load_workbook(patched_path, data_only=True)
    finally:
        os.remove(patched_path)

    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    records, unrecognized = flatten(rows)

    inv_lines = defaultdict(list)
    for r in records:
        inv_lines[r["DocNo"]].append(r)

    mismatches = []
    for docno, lines in inv_lines.items():
        total = sum(l["LineAmount"] for l in lines if isinstance(l["LineAmount"], (int, float)))
        inv_amt = lines[0]["InvoiceAmount"]
        if abs(total - inv_amt) > 0.05:
            mismatches.append((docno, total, inv_amt))

    return records, unrecognized, mismatches, len(inv_lines)


def convert_to_dataframe(input_path: str):
    """
    Same as convert_file(), but returns a pandas DataFrame instead of a
    list of dicts — this is what the combined pipeline uses to hand data
    from this stage straight to AutoCleaner in memory, no CSV round-trip.
    Returns (df, unrecognized, mismatches, invoice_count).
    """
    records, unrecognized, mismatches, invoice_count = convert_file(input_path)
    df = pd.DataFrame(records, columns=FIELDS)
    return df, unrecognized, mismatches, invoice_count
